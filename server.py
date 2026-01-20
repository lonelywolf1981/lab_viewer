from __future__ import annotations


def log_exception_to_file(prefix: str, exc: Exception) -> str:
    """Write traceback to a local file and return that file path."""
    import traceback as _tb
    try:
        base_dir = os.path.dirname(os.path.abspath(__file__))
        log_path = os.path.join(base_dir, "export_template_error.log")
        with open(log_path, "a", encoding="utf-8") as f:
            f.write("\n===== " + prefix + " " + dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S") + " =====\n")
            f.write(_tb.format_exc())
            f.write("\n")
        try:
            print("[TEMPLATE] ERROR logged to", log_path)
        except Exception:
            pass
        return log_path
    except Exception:
        return ""


# server.py
# Локальный веб-просмотрщик LeMuRe (Windows 7 friendly)
# Запуск: python server.py

import os
import re
import json
import io
import threading
import time as time_mod
import datetime as dt
import webbrowser
from bisect import bisect_left, bisect_right
from typing import List, Dict, Any, Tuple

from flask import Flask, render_template, request, jsonify, send_file


def _send_file_compat(fp, mimetype: str, filename: str):
    """send_file compat for different Flask versions (download_name vs attachment_filename)."""
    try:
        return send_file(fp, mimetype=mimetype, as_attachment=True, download_name=filename)
    except TypeError:
        return send_file(fp, mimetype=mimetype, as_attachment=True, attachment_filename=filename)


from lemure_reader import load_test, ChannelInfo
from functools import lru_cache
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor

executor = ThreadPoolExecutor(max_workers=2)

APP_HOST = "127.0.0.1"
APP_PORT = 8787

app = Flask(__name__)

STATE: Dict[str, Any] = {
    "loaded": False,
    "folder": "",
    "data": None,
    "t_list": [],
}

# Order persistence (saved in the same folder as server.py)
ORDER_FILE = os.path.join(os.path.dirname(__file__), "channel_order.json")
TEMPLATE_FILE = os.path.join(os.path.dirname(__file__), "template.xlsx")
ORDERS_DIR = os.path.join(os.path.dirname(__file__), 'saved_orders')
PRESETS_DIR = os.path.join(os.path.dirname(__file__), 'saved_presets')

# Optional viewer settings (editable without changing code)
# Настройки влияют на оформление экспорта «В шаблон XLSX».
# Их можно менять через UI (раздел «Оформление экспорта») или вручную в viewer_settings.json.
#
# Формат viewer_settings.json (канонический):
# {
#   "row_mark": {"threshold_T": 150, "color": "#FFF2CC", "intensity": 100},
#   "scales": {
#     "W": {"min": -1, "opt": 1, "colors": {"min": "#1CBCF2", "opt": "#00FF00", "max": "#F3919B"}},
#     "X": {"min": -1, "opt": 1, "colors": {"min": "#1CBCF2", "opt": "#00FF00", "max": "#F3919B"}},
#     "Y": {"min": -1, "opt": 1, "colors": {"min": "#1CBCF2", "opt": "#00FF00", "max": "#F3919B"}}
#   }
# }
# Для W/X/Y:
#   - min..opt (включая границы) = нормальный диапазон температуры
#   - ниже min  -> цвет colors.min
#   - в норме   -> цвет colors.opt
#   - выше opt  -> цвет colors.max
# intensity: 0..100 (0 = белый, 100 = выбранный цвет)
SETTINGS_FILE = os.path.join(os.path.dirname(__file__), 'viewer_settings.json')

DEFAULT_VIEWER_SETTINGS: Dict[str, Any] = {
    'row_mark': {
        'threshold_T': 150,
        'color': '#EAD706',  # мягкий жёлтый (как в примере)
        'intensity': 100,
    },
    'scales': {'W': {'min': -1, 'opt': 1, 'max': 2, 'colors': {'min': '#1CBCF2', 'opt': '#00FF00', 'max': '#F3919B'}},
               'X': {'min': -1, 'opt': 1, 'max': 2, 'colors': {'min': '#1CBCF2', 'opt': '#00FF00', 'max': '#F3919B'}},
               'Y': {'min': -1, 'opt': 1, 'max': 2, 'colors': {'min': '#1CBCF2', 'opt': '#00FF00', 'max': '#F3919B'}},
               },
}


def _deep_merge(dst: Dict[str, Any], src: Dict[str, Any]) -> Dict[str, Any]:
    for k, v in (src or {}).items():
        if isinstance(v, dict) and isinstance(dst.get(k), dict):
            _deep_merge(dst[k], v)
        else:
            dst[k] = v
    return dst


def _normalize_hex_color(s: str, default: str = '#FFF2CC') -> str:
    s = (s or '').strip()
    if not s:
        return default
    if not s.startswith('#'):
        s = '#' + s
    if re.fullmatch(r'#[0-9A-Fa-f]{6}', s):
        return s.upper()
    return default


def _argb_from_hex_and_intensity(hex_color: str, intensity_0_100: int) -> str:
    """Convert CSS #RRGGBB + intensity to Excel ARGB (FFRRGGBB), blending with white.
    intensity=100 -> original color; intensity=0 -> white.
    """
    hex_color = _normalize_hex_color(hex_color)
    i = max(0, min(100, int(intensity_0_100)))
    w = 1.0 - i / 100.0
    r = int(hex_color[1:3], 16)
    g = int(hex_color[3:5], 16)
    b = int(hex_color[5:7], 16)
    r2 = int(round(r * (1.0 - w) + 255 * w))
    g2 = int(round(g * (1.0 - w) + 255 * w))
    b2 = int(round(b * (1.0 - w) + 255 * w))
    return f'FF{r2:02X}{g2:02X}{b2:02X}'


def load_viewer_settings() -> Dict[str, Any]:
    # Start with defaults
    s = json.loads(json.dumps(DEFAULT_VIEWER_SETTINGS))

    user_s = {}
    try:
        if os.path.exists(SETTINGS_FILE):
            with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                user_s = json.load(f)
            if not isinstance(user_s, dict):
                user_s = {}
    except Exception:
        user_s = {}

    # Backward compatibility (older keys)
    # v7: yellow_threshold_T / yellow_intensity
    if isinstance(user_s, dict):
        if 'yellow_threshold_T' in user_s and 'row_mark' not in user_s:
            try:
                s['row_mark']['threshold_T'] = float(user_s.get('yellow_threshold_T'))
            except Exception:
                pass
        if 'yellow_intensity' in user_s and 'row_mark' not in user_s:
            try:
                s['row_mark']['intensity'] = int(user_s.get('yellow_intensity'))
            except Exception:
                pass

    _deep_merge(s, user_s)

    # Normalize
    rm = s.get('row_mark') if isinstance(s.get('row_mark'), dict) else {}
    try:
        rm['threshold_T'] = float(rm.get('threshold_T', 150))
    except Exception:
        rm['threshold_T'] = 150.0
    try:
        rm['intensity'] = max(0, min(100, int(rm.get('intensity', 100))))
    except Exception:
        rm['intensity'] = 100
    rm['color'] = _normalize_hex_color(str(rm.get('color') or ''), default='#FFF2CC')
    s['row_mark'] = rm

    scales = s.get('scales') if isinstance(s.get('scales'), dict) else {}
    out_scales = {}
    for k, defaults in DEFAULT_VIEWER_SETTINGS['scales'].items():
        d = scales.get(k, {}) if isinstance(scales.get(k), dict) else {}
        merged = dict(defaults)
        merged.update(d)
        for kk in ('min', 'opt', 'max'):
            try:
                merged[kk] = float(merged.get(kk))
            except Exception:
                merged[kk] = float(defaults[kk])
        # Guard against degenerate scales (must be increasing)
        if merged['min'] >= merged['opt']:
            merged['min'] = merged['opt'] - 1
        if merged['opt'] >= merged['max']:
            merged['max'] = merged['opt'] + 1
        # Colors for 3-point scale (min/opt/max)
        colors = merged.get('colors') if isinstance(merged.get('colors'), dict) else {}

        def _col(key, default):
            return _normalize_hex_color(str(colors.get(key) or ''), default=default)

        merged['colors'] = {
            'min': _col('min', '#0000FF'),
            'opt': _col('opt', '#00FF00'),
            'max': _col('max', '#FF0000'),
        }
        out_scales[k] = merged
    s['scales'] = out_scales

    return s


def normalize_viewer_settings(user_s: Dict[str, Any]) -> Dict[str, Any]:
    """Normalize/validate settings coming from UI/JSON."""
    s = json.loads(json.dumps(DEFAULT_VIEWER_SETTINGS))
    if not isinstance(user_s, dict):
        user_s = {}
    _deep_merge(s, user_s)

    rm = s.get('row_mark') if isinstance(s.get('row_mark'), dict) else {}
    try:
        rm['threshold_T'] = float(rm.get('threshold_T', 150))
    except Exception:
        rm['threshold_T'] = 150.0
    try:
        rm['intensity'] = max(0, min(100, int(rm.get('intensity', 100))))
    except Exception:
        rm['intensity'] = 100
    rm['color'] = _normalize_hex_color(str(rm.get('color') or ''), default='#FFF2CC')
    s['row_mark'] = rm

    scales = s.get('scales') if isinstance(s.get('scales'), dict) else {}
    out_scales = {}
    for k, defaults in DEFAULT_VIEWER_SETTINGS['scales'].items():
        d = scales.get(k, {}) if isinstance(scales.get(k), dict) else {}
        merged = dict(defaults)
        merged.update(d)
        for kk in ('min', 'opt', 'max'):
            try:
                merged[kk] = float(merged.get(kk))
            except Exception:
                merged[kk] = float(defaults[kk])
        if merged['min'] >= merged['opt']:
            merged['min'] = merged['opt'] - 1
        if merged['opt'] >= merged['max']:
            merged['max'] = merged['opt'] + 1
        # Colors for 3-point scale (min/opt/max)
        colors = merged.get('colors') if isinstance(merged.get('colors'), dict) else {}

        def _col(key, default):
            return _normalize_hex_color(str(colors.get(key) or ''), default=default)

        merged['colors'] = {
            'min': _col('min', '#0000FF'),
            'opt': _col('opt', '#00FF00'),
            'max': _col('max', '#FF0000'),
        }
        out_scales[k] = merged
    s['scales'] = out_scales
    return s


def save_viewer_settings(settings: Dict[str, Any]) -> None:
    try:
        with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
            json.dump(settings, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


VIEWER_SETTINGS = load_viewer_settings()


def _ensure_orders_dir() -> None:
    try:
        os.makedirs(ORDERS_DIR, exist_ok=True)
    except Exception:
        pass


def _sanitize_order_key(name: str) -> str:
    # Keep it filesystem-safe (Windows 7 friendly)
    name = (name or '').strip()
    name = re.sub(r'[^0-9A-Za-zА-Яа-я _.-]+', '_', name)
    name = re.sub(r'\s+', ' ', name).strip()
    name = name.replace('..', '.')
    if len(name) > 64:
        name = name[:64].strip()
    return name


def _order_path_by_key(key: str) -> str:
    _ensure_orders_dir()
    return os.path.join(ORDERS_DIR, f'{key}.json')


# Presets persistence
PRESETS_LOCK = threading.Lock()


def _ensure_presets_dir() -> None:
    try:
        os.makedirs(PRESETS_DIR, exist_ok=True)
    except Exception:
        pass


def _preset_path_by_key(key: str) -> str:
    _ensure_presets_dir()
    return os.path.join(PRESETS_DIR, f'{key}.json')


def load_saved_order() -> List[str]:
    try:
        if os.path.isfile(ORDER_FILE):
            import json
            with open(ORDER_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict) and isinstance(data.get("order"), list):
                return [str(x) for x in data["order"]]
            if isinstance(data, list):
                return [str(x) for x in data]
    except Exception:
        pass
    return []


def save_order(order: List[str]) -> None:
    try:
        import json
        with open(ORDER_FILE, "w", encoding="utf-8") as f:
            json.dump({"order": order}, f, ensure_ascii=False, indent=2)
    except Exception:
        # do not crash server
        pass


class ChannelResolver:
    """Fast O(1) channel lookup."""

    def __init__(self, cols: List[str]):
        self.cols = set(cols)
        self.by_suffix: Dict[str, List[str]] = {}
        for col in cols:
            if '-' in col:
                suffix = col.split('-', 1)[1]
                self.by_suffix.setdefault(suffix, []).append(col)

    def resolve(self, key: str, prefer: List[str] = None) -> str:
        if key in self.cols:
            return key
        candidates = self.by_suffix.get(key, [])
        if not candidates:
            return ""
        prefer = prefer or ["A-", "C-"]
        for prefix in prefer:
            for c in candidates:
                if c.startswith(prefix):
                    return c
        return candidates[0]


def _build_state(folder: str) -> Dict[str, Any]:
    data = load_test(folder)
    t_list = [r["t_ms"] for r in data["rows"]]
    return {"loaded": True, "folder": folder, "data": data, "t_list": t_list}


def _channel_to_dict(ch: ChannelInfo) -> Dict[str, str]:
    return {"code": ch.code, "name": ch.name, "unit": ch.unit, "label": ch.label}


def _summary(data: Dict[str, Any]) -> Dict[str, Any]:
    rows = data["rows"]
    if not rows:
        return {"points": 0}
    t0 = rows[0]["t_ms"]
    t1 = rows[-1]["t_ms"]
    return {
        "points": len(rows),
        "start_ms": t0,
        "end_ms": t1,
        "start": dt.datetime.fromtimestamp(t0 / 1000).strftime("%Y-%m-%d %H:%M:%S"),
        "end": dt.datetime.fromtimestamp(t1 / 1000).strftime("%Y-%m-%d %H:%M:%S"),
    }


def _slice_by_time(t_list: List[int], start_ms: int, end_ms: int) -> Tuple[int, int]:
    if start_ms > end_ms:
        start_ms, end_ms = end_ms, start_ms
    i0 = bisect_left(t_list, start_ms)
    i1 = bisect_right(t_list, end_ms)
    return i0, i1


@app.after_request
def add_no_cache_headers(resp):
    # Disable aggressive caching so browser always picks up latest JS/CSS after updates
    try:
        resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        resp.headers["Pragma"] = "no-cache"
        resp.headers["Expires"] = "0"
    except Exception:
        pass
    return resp


@app.route("/")
def index():
    # Версия статики: по mtime файлов, чтобы браузер гарантированно брал свежие JS/CSS.
    try:
        base = Path(__file__).resolve().parent
        mt = 0.0
        for p in [base / "static" / "app.js", base / "static" / "style.css", base / "templates" / "index.html"]:
            try:
                mt = max(mt, p.stat().st_mtime)
            except Exception:
                pass
        asset_v = str(int(mt))
    except Exception:
        asset_v = str(int(time_mod.time()))
    return render_template("index.html", port=APP_PORT, asset_v=asset_v)


@app.route("/favicon.ico")
def favicon():
    # Avoid 404 spam in run.log
    return ("", 204)


@app.route("/api/settings", methods=["GET", "POST"])
def api_settings():
    """Get or update viewer settings used for template export formatting.

    GET  -> {ok:true, settings:{...}}
    POST -> accepts either {settings:{...}} or just {...}
    """
    global VIEWER_SETTINGS
    if request.method == "GET":
        return jsonify({"ok": True, "settings": VIEWER_SETTINGS})

    body = request.get_json(force=True, silent=True) or {}
    user_s = body.get('settings') if isinstance(body, dict) and isinstance(body.get('settings'), dict) else body
    s = normalize_viewer_settings(user_s if isinstance(user_s, dict) else {})
    VIEWER_SETTINGS = s
    save_viewer_settings(s)
    return jsonify({"ok": True, "settings": s})


@app.route("/api/pick_folder", methods=["POST"])
def pick_folder():
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        root.wm_attributes("-topmost", 1)
        folder = filedialog.askdirectory(title="Выберите папку с тестом (где Prova*.dbf)")
        root.destroy()
        return jsonify({"ok": True, "folder": folder or ""})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e), "folder": ""})


def validate_folder_path(folder: str) -> bool:
    try:
        real_path = Path(folder).resolve()
        return real_path.exists() and real_path.is_dir()
    except Exception:
        return False


@app.route("/api/load", methods=["POST"])
def api_load():
    body = request.get_json(force=True, silent=True) or {}
    folder = (body.get("folder") or "").strip()

    if not validate_folder_path(folder):
        return jsonify({"ok": False, "error": "Недопустимый путь"})

    if not folder:
        return jsonify({"ok": False, "error": "Путь к папке пустой"})

    if not os.path.isdir(folder):
        return jsonify({"ok": False, "error": "Папка не существует: " + folder})

    try:
        st = _build_state(folder)
        STATE.update(st)
        try:
            _cached_series_slice.cache_clear()
        except Exception:
            pass
        data = STATE["data"]
        channels: Dict[str, ChannelInfo] = data["channels"]
        cols = data["cols"]

        ch_list = []
        for c in cols:
            if c in channels:
                ch_list.append(_channel_to_dict(channels[c]))
            else:
                ch_list.append({"code": c, "name": "", "unit": "", "label": c})

        return jsonify({
            "ok": True,
            "folder": data["root"],
            "meta": data["meta"],
            "summary": _summary(data),
            "channels": ch_list,
            "file_order": [c.get("code") for c in ch_list],
            "saved_order": load_saved_order(),
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@lru_cache(maxsize=8)
def _cached_series_slice(data_id: int, channels_key: str, start_ms: int, end_ms: int, step_i: int):
    """Return (t_ms, series_dict) for the requested channels and time range.

    Cached to speed up repeated redraws with the same parameters.
    """
    rows = STATE["data"]["rows"]
    t_list = STATE["t_list"]
    channels = [c.strip() for c in channels_key.split(",") if c.strip()]

    i0, i1 = _slice_by_time(t_list, start_ms, end_ms)
    sliced = rows[i0:i1:step_i]

    t = [r["t_ms"] for r in sliced]

    def _to_float(v):
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return float(v)
        try:
            return float(str(v).replace(",", "."))
        except Exception:
            return None

    series = {}
    for code in channels:
        series[code] = [_to_float(r.get(code)) for r in sliced]

    return t, series


@app.route("/api/series", methods=["GET"])
def api_series():
    if not STATE.get("loaded"):
        return jsonify({"ok": False, "error": "Данные не загружены"})

    data = STATE["data"]
    rows: List[Dict[str, Any]] = data["rows"]
    t_list: List[int] = STATE["t_list"]

    channels = request.args.get("channels", "")
    ch = [c for c in channels.split(",") if c.strip()]
    if not ch:
        return jsonify({"ok": False, "error": "Не выбраны каналы"})

    try:
        start_ms = int(float(request.args.get("start_ms", rows[0]["t_ms"])))
        end_ms = int(float(request.args.get("end_ms", rows[-1]["t_ms"])))
    except Exception:
        start_ms = rows[0]["t_ms"]
        end_ms = rows[-1]["t_ms"]

    if start_ms > end_ms:
        start_ms, end_ms = end_ms, start_ms

    # Manual step (default)
    try:
        step_i = max(1, int(request.args.get("step", "1")))
    except Exception:
        step_i = 1

    # Auto: ask server to keep around N points
    max_points = request.args.get("max_points", "")
    try:
        target = int(float(max_points)) if str(max_points).strip() != "" else 0
    except Exception:
        target = 0

    # Determine slice once (also used for cache heuristics)
    i0, i1 = _slice_by_time(t_list, start_ms, end_ms)
    raw_pts = max(0, i1 - i0)

    if target and target > 0:
        # ceil(raw_pts / target)
        step_i = max(1, (raw_pts + target - 1) // target)

    # Heuristic: avoid caching too large responses (can waste RAM)
    pts_after = max(0, (raw_pts + step_i - 1) // step_i) if raw_pts > 0 else 0
    cells = pts_after * len(ch)
    use_cache = (pts_after <= 40000) and (cells <= 400000)

    channels_key = ",".join(ch)

    try:
        if use_cache:
            t_ms, series = _cached_series_slice(
                id(rows),
                channels_key,
                start_ms,
                end_ms,
                step_i,
            )
        else:
            # Uncached path for very large payloads
            sliced = rows[i0:i1:step_i]
            t_ms = [r["t_ms"] for r in sliced]

            def _to_float(v):
                if v is None:
                    return None
                if isinstance(v, (int, float)):
                    return float(v)
                try:
                    return float(str(v).replace(",", "."))
                except Exception:
                    return None

            series = {code: [_to_float(r.get(code)) for r in sliced] for code in ch}

        return jsonify({
            "ok": True,
            "t_ms": t_ms,
            "series": series,
            "step": step_i,
            "points": len(t_ms),
        })

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/range_stats", methods=["GET"])
def api_range_stats():
    """Return number of raw points within [start_ms..end_ms].

    Used by UI for accurate auto-step and export info.
    """
    if not STATE.get("loaded"):
        return jsonify({"ok": False, "error": "Данные не загружены"}), 400

    t_list = STATE.get("t_list") or []
    if not t_list:
        return jsonify({"ok": True, "points": 0, "total": 0})

    try:
        start_ms = int(float(request.args.get("start_ms", t_list[0])))
        end_ms = int(float(request.args.get("end_ms", t_list[-1])))
    except Exception:
        start_ms = t_list[0]
        end_ms = t_list[-1]

    if start_ms > end_ms:
        start_ms, end_ms = end_ms, start_ms

    i0, i1 = _slice_by_time(t_list, start_ms, end_ms)
    pts = max(0, i1 - i0)
    return jsonify({
        "ok": True,
        "start_ms": start_ms,
        "end_ms": end_ms,
        "points": pts,
        "total": len(t_list),
    })


@app.route("/api/save_order", methods=["POST"])
def api_save_order():
    body = request.get_json(force=True, silent=True) or {}
    order = body.get("order")
    if not isinstance(order, list):
        return jsonify({"ok": False, "error": "order must be a list"}), 400
    seen = set()
    normalized = []
    for x in order:
        s = str(x)
        if s not in seen:
            seen.add(s)
            normalized.append(s)
    save_order(normalized)
    return jsonify({"ok": True, "saved": len(normalized)})


@app.route("/api/orders_list", methods=["GET"])
def api_orders_list():
    """Return list of saved named orders."""
    try:
        _ensure_orders_dir()
        items = []
        for fn in os.listdir(ORDERS_DIR):
            if not fn.lower().endswith('.json'):
                continue
            path = os.path.join(ORDERS_DIR, fn)
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                key = str(data.get('key') or os.path.splitext(fn)[0])
                name = str(data.get('name') or key)
                order = data.get('order')
                count = len(order) if isinstance(order, list) else 0
                saved_at = str(data.get('saved_at') or '')
                items.append({"key": key, "name": name, "count": count, "saved_at": saved_at})
            except Exception:
                # ignore broken files
                continue
        items.sort(key=lambda x: x.get('name', '').lower())
        return jsonify({"ok": True, "orders": items})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e), "orders": []})


@app.route("/api/orders_save", methods=["POST"])
def api_orders_save():
    """Save a named order."""
    body = request.get_json(force=True, silent=True) or {}
    name = str(body.get('name') or '').strip()
    order = body.get('order')
    if not name:
        return jsonify({"ok": False, "error": "Имя не задано"}), 400
    if not isinstance(order, list):
        return jsonify({"ok": False, "error": "order must be a list"}), 400

    key = _sanitize_order_key(name)
    if not key:
        return jsonify({"ok": False, "error": "Имя некорректное"}), 400

    seen = set()
    normalized = []
    for x in order:
        s = str(x)
        if s and s not in seen:
            seen.add(s)
            normalized.append(s)

    payload = {
        "name": name,
        "key": key,
        "order": normalized,
        "saved_at": dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }
    try:
        path = _order_path_by_key(key)
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        return jsonify({"ok": True, "key": key, "saved": len(normalized)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/orders_load", methods=["GET"])
def api_orders_load():
    """Load a named order by key (or by name)."""
    key = str(request.args.get('key') or '').strip()
    name = str(request.args.get('name') or '').strip()
    if not key and name:
        key = _sanitize_order_key(name)
    if not key:
        return jsonify({"ok": False, "error": "Не задан ключ (key)"}), 400

    try:
        path = _order_path_by_key(key)
        if not os.path.isfile(path):
            return jsonify({"ok": False, "error": "Сохранённый порядок не найден: " + key}), 404
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        order = data.get('order')
        if not isinstance(order, list):
            order = []
        # normalize
        out = []
        seen = set()
        for x in order:
            s = str(x)
            if s and s not in seen:
                seen.add(s)
                out.append(s)
        return jsonify({"ok": True, "key": key, "name": str(data.get('name') or key), "order": out})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/presets_list", methods=["GET"])
def api_presets_list():
    """Return list of saved presets."""
    try:
        _ensure_presets_dir()
        items = []
        for fn in os.listdir(PRESETS_DIR):
            if not fn.lower().endswith('.json'):
                continue
            path = os.path.join(PRESETS_DIR, fn)
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                key = str(data.get('key') or os.path.splitext(fn)[0])
                name = str(data.get('name') or key)
                channels = data.get('channels')
                if not isinstance(channels, list):
                    channels = (data.get('preset') or {}).get('channels') if isinstance(data.get('preset'),
                                                                                        dict) else []
                count = len(channels) if isinstance(channels, list) else 0
                saved_at = str(data.get('saved_at') or '')
                items.append({"key": key, "name": name, "count": count, "saved_at": saved_at})
            except Exception:
                continue
        items.sort(key=lambda x: x.get('name', '').lower())
        return jsonify({"ok": True, "presets": items})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e), "presets": []})


@app.route("/api/presets_save", methods=["POST"])
def api_presets_save():
    """Save a named preset (channels + settings)."""
    body = request.get_json(force=True, silent=True) or {}
    name = str(body.get('name') or '').strip()
    preset = body.get('preset') if isinstance(body.get('preset'), dict) else {}
    if not name:
        return jsonify({"ok": False, "error": "Имя не задано"}), 400

    key = _sanitize_order_key(name)
    if not key:
        return jsonify({"ok": False, "error": "Имя некорректное"}), 400

    # Normalize preset structure
    channels = preset.get('channels')
    if not isinstance(channels, list):
        channels = []
    channels_n = []
    seen = set()
    for x in channels:
        s = str(x)
        if s and s not in seen:
            seen.add(s)
            channels_n.append(s)

    order = preset.get('order')
    if isinstance(order, list):
        o2 = []
        seen2 = set()
        for x in order:
            s = str(x)
            if s and s not in seen2:
                seen2.add(s)
                o2.append(s)
        order = o2
    else:
        order = []

    payload = {
        "name": name,
        "key": key,
        "saved_at": dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        "channels": channels_n,
        "sort_mode": str(preset.get('sort_mode') or ''),
        "order": order,
        "step_auto": bool(preset.get('step_auto')) if isinstance(preset.get('step_auto'), bool) else True,
        "step_target": int(preset.get('step_target') or 5000),
        "step": int(preset.get('step') or 1),
        "show_legend": bool(preset.get('show_legend')) if isinstance(preset.get('show_legend'), bool) else True,
    }

    try:
        _ensure_presets_dir()
        path = _preset_path_by_key(key)
        with PRESETS_LOCK:
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(payload, f, ensure_ascii=False, indent=2)
        return jsonify({"ok": True, "key": key, "count": len(channels_n)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/presets_load", methods=["GET"])
def api_presets_load():
    """Load preset by key (or by name)."""
    key = str(request.args.get('key') or '').strip()
    name = str(request.args.get('name') or '').strip()
    if not key and name:
        key = _sanitize_order_key(name)
    if not key:
        return jsonify({"ok": False, "error": "Не задан ключ (key)"}), 400

    try:
        path = _preset_path_by_key(key)
        if not os.path.isfile(path):
            return jsonify({"ok": False, "error": "Сохранённый набор не найден: " + key}), 404
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        # Backward/forward compatibility: return full object, plus 'preset' alias
        preset = dict(data)
        preset.pop('ok', None)
        return jsonify({"ok": True, "key": key, "name": str(data.get('name') or key), "preset": preset})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/presets_delete", methods=["POST"])
def api_presets_delete():
    body = request.get_json(force=True, silent=True) or {}
    key = str(body.get('key') or '').strip()
    if not key:
        return jsonify({"ok": False, "error": "Не задан key"}), 400
    try:
        path = _preset_path_by_key(key)
        if not os.path.isfile(path):
            return jsonify({"ok": False, "error": "Набор не найден: " + key}), 404
        with PRESETS_LOCK:
            os.remove(path)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


def _export_csv(rows: List[Dict[str, Any]], channels: List[str]) -> bytes:
    import csv
    out = io.StringIO()
    writer = csv.writer(out, delimiter=";")
    writer.writerow(["timestamp"] + channels)
    for r in rows:
        ts = dt.datetime.fromtimestamp(r["t_ms"] / 1000).strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
        writer.writerow([ts] + [r.get(c) if r.get(c) is not None else "" for c in channels])
    return out.getvalue().encode("utf-8-sig")


def _export_xlsx(rows: List[Dict[str, Any]], channels: List[str]) -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    ws.append(["timestamp"] + channels)
    for r in rows:
        ts = dt.datetime.fromtimestamp(r["t_ms"] / 1000).strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
        ws.append([ts] + [r.get(c) for c in channels])
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


@app.route("/api/export", methods=["GET"])
def api_export():
    if not STATE.get("loaded"):
        return jsonify({"ok": False, "error": "Данные не загружены"}), 400

    data = STATE["data"]
    rows: List[Dict[str, Any]] = data["rows"]
    t_list: List[int] = STATE["t_list"]

    fmt = (request.args.get("format", "csv") or "csv").lower()
    channels = request.args.get("channels", "")
    ch = [c for c in channels.split(",") if c.strip()]
    if not ch:
        return jsonify({"ok": False, "error": "Не выбраны каналы"}), 400

    try:
        start_ms = int(float(request.args.get("start_ms", rows[0]["t_ms"])))
        end_ms = int(float(request.args.get("end_ms", rows[-1]["t_ms"])))
    except Exception:
        start_ms = rows[0]["t_ms"]
        end_ms = rows[-1]["t_ms"]

    try:
        step_i = max(1, int(request.args.get("step", "1")))
    except Exception:
        step_i = 1

    i0, i1 = _slice_by_time(t_list, start_ms, end_ms)
    sliced = rows[i0:i1:step_i]

    if fmt == "xlsx":
        payload = _export_xlsx(sliced, ch)
        return _send_file_compat(io.BytesIO(payload),
                                 "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 "export.xlsx")
    else:
        payload = _export_csv(sliced, ch)
        return _send_file_compat(io.BytesIO(payload),
                                 "text/csv",
                                 "export.csv")


def _nearest_index(t_list: List[int], target_ms: int) -> int:
    """Return nearest index in sorted t_list to target_ms."""
    if not t_list:
        return -1
    i = bisect_left(t_list, target_ms)
    if i <= 0:
        return 0
    if i >= len(t_list):
        return len(t_list) - 1
    a = t_list[i - 1]
    b = t_list[i]
    if abs(target_ms - a) <= abs(b - target_ms):
        return i - 1
    return i


def _resolve_channel(cols: List[str], key: str) -> str:
    """Resolve sensor code from available columns by semantic key."""
    # special channels without prefix
    if key in ("T-sie", "UR-sie"):
        return key if key in cols else ""

    # common suffix mapping (prefer A- then C-)
    candidates = [f"A-{key}", f"C-{key}", key]
    for c in candidates:
        if c in cols:
            return c

    # fallback: find any ending with -key
    suf = "-" + key
    for c in cols:
        if c.endswith(suf):
            return c
    return ""


# ---------------- Template export (fill template.xlsx preserving formatting) ----------------

@app.route("/api/export_template", methods=["GET"])
def api_export_template():
    try:
        return _api_export_template_impl()
    except Exception as e:
        lp = log_exception_to_file("api_export_template", e)
        msg = str(e)
        if lp:
            msg = msg + f" (подробности в {os.path.basename(lp)})"
        return jsonify({"ok": False, "error": msg}), 500


def _api_export_template_impl():
    """Заполнить template.xlsx, сохранив форматирование/цвета/формулы.

    Правила заполнения (как вы описали):
      - Строки данных начинаются с 4-й строки.
      - Колонка B: время испытания = 00:00:00 + N*20 секунд.
      - Колонка C: текущее время берём из файла (только время суток, без даты).
      - Колонки D..T: сопоставляем с показаниями датчиков.

    Важно: в шаблоне есть формулы (например U..Y). Их сохраняем и протягиваем вниз,
    а значения в остальных «сырых» колонках очищаем, чтобы не оставались примерные данные.
    """

    if not STATE.get("loaded"):
        return jsonify({"ok": False, "error": "Данные не загружены"}), 400

    data = STATE["data"]
    rows: List[Dict[str, Any]] = data["rows"]
    t_list: List[int] = STATE["t_list"]
    cols: List[str] = data.get("cols") or []
    channels: Dict[str, ChannelInfo] = data.get("channels") or {}

    if not rows:
        return jsonify({"ok": False, "error": "Нет строк данных"}), 400

    try:
        start_ms = int(float(request.args.get("start_ms", rows[0]["t_ms"])))
        end_ms = int(float(request.args.get("end_ms", rows[-1]["t_ms"])))
    except Exception:
        start_ms = rows[0]["t_ms"]
        end_ms = rows[-1]["t_ms"]

    if start_ms > end_ms:
        start_ms, end_ms = end_ms, start_ms

    i0 = bisect_left(t_list, start_ms)
    if i0 >= len(t_list):
        return jsonify({"ok": False, "error": "Диапазон вне данных"}), 400
    t0 = t_list[i0]
    if t0 > end_ms:
        return jsonify({"ok": False, "error": "Пустой диапазон"}), 400

    # Построение сетки 20s
    grid_ms: List[int] = []
    g = t0
    step_ms = 20000
    while g <= end_ms:
        grid_ms.append(g)
        g += step_ms

    idxs: List[int] = []
    for g in grid_ms:
        idx = _nearest_index(t_list, g)
        if idx < 0 or abs(t_list[idx] - g) > 30000:
            idxs.append(-1)
        else:
            idxs.append(idx)

    ch_arg = (request.args.get("channels") or "").strip()
    selected_list = [c.strip() for c in ch_arg.split(",") if c.strip()]
    selected = set(selected_list)

    def resolve_for_selection(key: str) -> str:
        matches = []
        if key in cols:
            matches.append(key)
        suf = f"-{key}"
        for c in cols:
            if c.endswith(suf) and c not in matches:
                matches.append(c)
        if not matches:
            return ""
        if selected:
            for pref in ("A-", "C-"):
                for c in matches:
                    if c in selected and c.startswith(pref):
                        return c
            for c in matches:
                if c in selected:
                    return c
            return ""
        for pref in ("A-", "C-"):
            for c in matches:
                if c.startswith(pref):
                    return c
        return matches[0]

    key_to_col = {
        "Pc": 4, "Pe": 5, "T-sie": 6, "UR-sie": 7,
        "Tc": 8, "Te": 9, "T1": 10, "T2": 11,
        "T3": 12, "T4": 13, "T5": 14, "T6": 15,
        "T7": 16, "I": 17, "F": 18, "V": 19, "W": 20,
    }

    key_to_code = {k: resolve_for_selection(k) for k in key_to_col.keys()}
    fixed_codes = set([v for v in key_to_code.values() if v])

    candidate_codes = [c for c in selected_list if c in cols] if selected_list else list(cols)

    try:
        include_extra = int(str(request.args.get("include_extra", "1")).strip() or "1")
    except Exception:
        include_extra = 1

    extra_codes = [c for c in candidate_codes if c and c not in fixed_codes]
    if include_extra <= 0:
        extra_codes = []

    def _nat_key(s: str):
        import re as _re
        parts = _re.split(r'(\d+)', s or '')
        out = []
        for part in parts:
            if part.isdigit():
                try:
                    out.append(int(part))
                except Exception:
                    out.append(part)
            else:
                out.append(part.casefold())
        return out

    def _display_name(code: str) -> str:
        ch = channels.get(code)
        nm = ''
        try:
            nm = (ch.name or '').strip() if ch else ''
        except Exception:
            nm = ''
        if nm:
            return nm
        lb = ''
        try:
            lb = (ch.label or '').strip() if ch else ''
        except Exception:
            lb = ''
        return lb or code

    def _template_header(code: str) -> str:
        ch = channels.get(code)
        nm = _display_name(code)
        unit = ''
        try:
            unit = (ch.unit or '').strip() if ch else ''
        except Exception:
            unit = ''
        if unit:
            return f"{nm} [{unit}]"
        return nm

    extra_codes.sort(key=lambda c: (_nat_key(_display_name(c)), _nat_key(c)))

    EXTRA_START_COL = 26
    extra_col_map = {code: (EXTRA_START_COL + i) for i, code in enumerate(extra_codes)}

    template_path = TEMPLATE_FILE
    if not os.path.isfile(template_path):
        return jsonify({"ok": False, "error": "Не найден template.xlsx"}), 400

    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.formula.translate import Translator
        from copy import copy
        import datetime as _dt

        t_all0 = time_mod.perf_counter()

        # Загружаем шаблон (data_only=False: формулы сохраняем)
        wb = load_workbook(template_path, data_only=False, keep_vba=False)
        ws = wb.active

        # Хладагент (ячейка B1)
        refrigerant = (request.args.get('refrigerant') or 'R290').strip()
        if refrigerant not in ('R290', 'R600a'):
            refrigerant = 'R290'
        try:
            ws['B1'].value = refrigerant
        except Exception:
            pass

        # Отключаем автопересчёт формул (ускоряет заполнение/сохранение)
        try:
            wb.calculation.calcMode = 'manual'
        except Exception:
            pass

        t_load = time_mod.perf_counter()

        start_row = 4
        pattern_row = start_row

        required_last_col = 0
        if extra_codes:
            required_last_col = EXTRA_START_COL + len(extra_codes) - 1
        BASE_LAST_COL = 25
        max_col = max(BASE_LAST_COL, required_last_col or 0)

        # Формульные колонки + исходные формулы из строки-шаблона
        formula_cols = set()
        pattern_formulas = {}
        for c in range(1, max_col + 1):
            v = ws.cell(row=pattern_row, column=c).value
            if isinstance(v, str) and v.startswith('='):
                formula_cols.add(c)
                pattern_formulas[c] = v
        formula_list = sorted(list(formula_cols))

        # Готовим Translator один раз на колонку (вместо пересоздания на каждую строку)
        translators = {}
        for c, src_formula in pattern_formulas.items():
            src_addr = f"{get_column_letter(c)}{pattern_row}"
            try:
                translators[c] = Translator(src_formula, origin=src_addr)
            except Exception:
                translators[c] = None

        # Какие колонки реально заполняем значениями
        writers = []
        for key, colnum in key_to_col.items():
            code = key_to_code.get(key) or ""
            if not code:
                continue
            if selected and code not in selected:
                continue
            writers.append((code, colnum))

        extra_writers = []
        if extra_codes:
            for code in extra_codes:
                colnum = extra_col_map.get(code)
                if colnum:
                    extra_writers.append((code, colnum))

        # Базовые колонки D..T (чтобы гарантированно очищать «невыбранные» значения)
        base_raw_cols = sorted(set(key_to_col.values()))

        needed_last_row = start_row + len(idxs) - 1

        # Ключевая оптимизация:
        # 1) Не трогаем стили/формат на строках, которые уже есть в template.xlsx.
        # 2) Стили копируем ТОЛЬКО если нужно создать строки ниже максимума шаблона.
        template_max_row = ws.max_row
        if needed_last_row > template_max_row:
            # Кэшируем только то, что реально влияет на формат (стиль + number_format)
            style_cache = {}
            for c in range(1, max_col + 1):
                src = ws.cell(row=pattern_row, column=c)
                try:
                    style_cache[c] = (copy(src._style), src.number_format)
                except Exception:
                    style_cache[c] = (None, src.number_format)

            pat_height = None
            try:
                pat_height = ws.row_dimensions[pattern_row].height
            except Exception:
                pat_height = None

            for r in range(template_max_row + 1, needed_last_row + 1):
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    st, nf = style_cache.get(c, (None, None))
                    if st is not None:
                        cell._style = st
                    if nf is not None:
                        cell.number_format = nf
                if pat_height is not None:
                    try:
                        ws.row_dimensions[r].height = pat_height
                    except Exception:
                        pass

        t_prep = time_mod.perf_counter()

        # Заполняем значения (только нужные колонки), не создавая лишних «пустых» ячеек.
        for j, idx in enumerate(idxs):
            r = start_row + j

            # B: смещение времени теста
            ws.cell(row=r, column=2).value = _dt.timedelta(seconds=j * 20)

            # C: время (time)
            if idx >= 0:
                tms = rows[idx]["t_ms"]
                dt_obj = dt.datetime.fromtimestamp(tms / 1000)
                ws.cell(row=r, column=3).value = dt_obj.time()
            else:
                ws.cell(row=r, column=3).value = None

            # Сначала гарантированно очищаем базовые колонки D..T,
            # чтобы в строках шаблона не осталось старых значений.
            for colnum in base_raw_cols:
                ws.cell(row=r, column=colnum).value = None

            if idx >= 0:
                row = rows[idx]

                # Основные колонки (D..T)
                for code, colnum in writers:
                    v = row.get(code)
                    if v is not None:
                        try:
                            ws.cell(row=r, column=colnum).value = float(v)
                        except Exception:
                            pass

                # Доп. сенсоры (Z…)
                for code, colnum in extra_writers:
                    v = row.get(code)
                    if v is None:
                        ws.cell(row=r, column=colnum).value = None
                    else:
                        try:
                            ws.cell(row=r, column=colnum).value = float(v)
                        except Exception:
                            ws.cell(row=r, column=colnum).value = None
            else:
                # Если данных нет — обязательно очистим extra колонки
                for _, colnum in extra_writers:
                    ws.cell(row=r, column=colnum).value = None

            # Формулы: ставим только если в этой строке они отсутствуют
            for c in formula_list:
                cell = ws.cell(row=r, column=c)
                vv = cell.value
                if isinstance(vv, str) and vv.startswith('='):
                    continue

                src_formula = pattern_formulas.get(c)
                if not src_formula:
                    continue

                tr = translators.get(c)
                if tr is not None:
                    dst_addr = f"{get_column_letter(c)}{r}"
                    try:
                        cell.value = tr.translate_formula(dst_addr)
                    except Exception:
                        cell.value = src_formula
                else:
                    cell.value = src_formula

        # Заголовки для extra сенсоров
        if extra_writers:
            header_row = start_row - 1
            for code, colnum in extra_writers:
                try:
                    ws.cell(row=header_row, column=colnum).value = _template_header(code)
                except Exception:
                    ws.cell(row=header_row, column=colnum).value = _display_name(code)

        # Очистка строк ниже (быстро: не создаём новые ячейки)
        clear_from = start_row + len(idxs)
        clear_to = min(ws.max_row, clear_from + 200)
        clear_cols = [2, 3] + base_raw_cols + [c for _, c in extra_writers]
        cells_dict = getattr(ws, '_cells', None)

        for r in range(clear_from, clear_to + 1):
            for c in clear_cols:
                if cells_dict is not None:
                    cell = cells_dict.get((r, c))
                    if cell is not None:
                        cell.value = None
                else:
                    ws.cell(row=r, column=c).value = None

        t_write = time_mod.perf_counter()

        # Условное форматирование (как было)
        try:
            from openpyxl.styles import PatternFill
            from openpyxl.formatting.rule import FormulaRule

            first_r = start_row
            last_r = start_row + len(idxs) - 1

            if last_r >= first_r:
                rm = VIEWER_SETTINGS.get('row_mark') if isinstance(VIEWER_SETTINGS.get('row_mark'), dict) else {}
                thr = rm.get('threshold_T', 150)
                try:
                    thr_f = float(thr)
                    thr = int(thr_f) if thr_f.is_integer() else thr_f
                except Exception:
                    thr = 150

                color_hex = str(rm.get('color') or '#EAD706')
                intensity = rm.get('intensity', 100)
                try:
                    intensity = int(intensity)
                except Exception:
                    intensity = 100

                fill_argb = _argb_from_hex_and_intensity(color_hex, intensity)
                row_fill = PatternFill(fill_type='solid', start_color=fill_argb, end_color=fill_argb)

                rng_row = f"B{first_r}:P{last_r}"
                rule_row = FormulaRule(formula=[f"$T{first_r}<{thr}"], fill=row_fill, stopIfTrue=True)
                try:
                    rule_row.priority = 1
                except Exception:
                    pass
                ws.conditional_formatting.add(rng_row, rule_row)

                scales = VIEWER_SETTINGS.get('scales') if isinstance(VIEWER_SETTINGS.get('scales'), dict) else {}

                def _hex_to_argb(hx: str, default: str) -> str:
                    hx = _normalize_hex_color(str(hx or ''), default=default)
                    return 'FF' + hx[1:].upper()

                def _fmt_num(x) -> str:
                    try:
                        xf = float(x)
                        if xf.is_integer():
                            return str(int(xf))
                        return (f"{xf:g}").replace(',', '.')
                    except Exception:
                        return str(x)

                def _add_discrete(col_letter: str, base_prio: int) -> None:
                    spec = scales.get(col_letter) if isinstance(scales.get(col_letter), dict) else {}
                    dflt = DEFAULT_VIEWER_SETTINGS['scales'][col_letter]

                    vmin = spec.get('min', dflt.get('min'))
                    vopt = spec.get('opt', dflt.get('opt'))

                    try:
                        vmin_f = float(vmin)
                    except Exception:
                        vmin_f = float(dflt.get('min') or 0)
                    try:
                        vopt_f = float(vopt)
                    except Exception:
                        vopt_f = float(dflt.get('opt') or 0)

                    if vmin_f > vopt_f:
                        vmin_f, vopt_f = vopt_f, vmin_f

                    min_s = _fmt_num(vmin_f)
                    opt_s = _fmt_num(vopt_f)

                    colors = spec.get('colors') if isinstance(spec.get('colors'), dict) else {}
                    dcolors = dflt.get('colors') if isinstance(dflt.get('colors'), dict) else {}

                    c_lo = _hex_to_argb(colors.get('min'), dcolors.get('min', '#1CBCF2'))
                    c_mid = _hex_to_argb(colors.get('opt'), dcolors.get('opt', '#00FF00'))
                    c_hi = _hex_to_argb(colors.get('max'), dcolors.get('max', '#F3919B'))

                    f_lo = PatternFill(fill_type='solid', start_color=c_lo, end_color=c_lo)
                    f_mid = PatternFill(fill_type='solid', start_color=c_mid, end_color=c_mid)
                    f_hi = PatternFill(fill_type='solid', start_color=c_hi, end_color=c_hi)

                    rng = f"{col_letter}{first_r}:{col_letter}{last_r}"

                    r1 = FormulaRule(
                        formula=[f'AND(${col_letter}{first_r}<>"",${col_letter}{first_r}<{min_s})'],
                        fill=f_lo, stopIfTrue=True)
                    r2 = FormulaRule(
                        formula=[
                            f'AND(${col_letter}{first_r}<>"",${col_letter}{first_r}>={min_s},${col_letter}{first_r}<={opt_s})'],
                        fill=f_mid, stopIfTrue=True)
                    r3 = FormulaRule(
                        formula=[f'AND(${col_letter}{first_r}<>"",${col_letter}{first_r}>{opt_s})'],
                        fill=f_hi, stopIfTrue=True)

                    try:
                        r1.priority = base_prio
                        r2.priority = base_prio + 1
                        r3.priority = base_prio + 2
                    except Exception:
                        pass

                    ws.conditional_formatting.add(rng, r1)
                    ws.conditional_formatting.add(rng, r2)
                    ws.conditional_formatting.add(rng, r3)

                _add_discrete('W', 10)
                _add_discrete('X', 20)
                _add_discrete('Y', 30)

        except Exception as _cf_e:
            try:
                print("[TEMPLATE] conditional formatting skipped:", _cf_e)
            except Exception:
                pass

        t_before_save = time_mod.perf_counter()

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        t_after_save = time_mod.perf_counter()

        total_s = t_after_save - t_all0
        load_s = t_load - t_all0
        prep_s = t_prep - t_load
        write_s = t_write - t_prep
        format_s = t_before_save - t_write
        save_s = t_after_save - t_before_save

        try:
            print(
                f"[TEMPLATE] timing: load={load_s:.3f}s prep={prep_s:.3f}s write={write_s:.3f}s format={format_s:.3f}s save={save_s:.3f}s total={total_s:.3f}s rows={len(idxs)} max_col={max_col}")
        except Exception:
            pass

        fn = f"template_filled_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        resp = _send_file_compat(bio, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", fn)

        try:
            resp.headers['X-Export-Total-S'] = f"{total_s:.3f}"
            resp.headers['X-Export-Timing'] = f"load {load_s:.1f}s | prep {prep_s:.1f}s | write {write_s:.1f}s | format {format_s:.1f}s | save {save_s:.1f}s | total {total_s:.1f}s"
        except Exception:
            pass

        return resp

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

def _open_browser():
    try:
        webbrowser.open(f"http://{APP_HOST}:{APP_PORT}/", new=1)
    except Exception:
        pass


if __name__ == "__main__":
    threading.Timer(0.8, _open_browser).start()
    _ensure_orders_dir()
    _ensure_presets_dir()
    # threaded=False — чтобы tkinter "Обзор..." работал предсказуемо
    # app.run(host=APP_HOST, port=APP_PORT, debug=False, threaded=False)
    app.run(host=APP_HOST, port=APP_PORT, debug=False, threaded=True, use_reloader=False)
