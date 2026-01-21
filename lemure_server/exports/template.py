from __future__ import annotations

import io
import os
import time as time_mod
import datetime as dt
from bisect import bisect_left
from typing import Any, Dict, List

from flask import jsonify

from lemure_reader import ChannelInfo

from ..config import TEMPLATE_FILE, PROJECT_ROOT
from ..state import STATE, nearest_index
from ..settings import (
    get_viewer_settings,
    DEFAULT_VIEWER_SETTINGS,
    _argb_from_hex_and_intensity,
    _normalize_hex_color,
)


def export_template_impl(args) -> Any:
    """Fill template.xlsx preserving formatting/colors/formulas."""

    if not STATE.get("loaded"):
        return jsonify({"ok": False, "error": "Данные не загружены"}), 400

    data = STATE["data"]
    rows: List[Dict[str, Any]] = data["rows"]
    t_list: List[int] = STATE["t_list"]
    cols: List[str] = data.get("cols") or []
    channels: Dict[str, ChannelInfo] = data.get("channels") or {}

    if not rows:
        return jsonify({"ok": False, "error": "Нет строк данных"}), 400

    try:
        start_ms = int(float(args.get("start_ms", rows[0]["t_ms"])))
        end_ms = int(float(args.get("end_ms", rows[-1]["t_ms"])))
    except Exception:
        start_ms = rows[0]["t_ms"]
        end_ms = rows[-1]["t_ms"]

    if start_ms > end_ms:
        start_ms, end_ms = end_ms, start_ms

    i0 = bisect_left(t_list, start_ms)
    if i0 >= len(t_list):
        return jsonify({"ok": False, "error": "Диапазон вне данных"}), 400
    t0 = t_list[i0]
    if t0 > end_ms:
        return jsonify({"ok": False, "error": "Пустой диапазон"}), 400

    # build 20s grid
    grid_ms: List[int] = []
    g = t0
    step_ms = 20000
    while g <= end_ms:
        grid_ms.append(g)
        g += step_ms

    idxs: List[int] = []
    for g in grid_ms:
        idx = nearest_index(t_list, g)
        if idx < 0 or abs(t_list[idx] - g) > 30000:
            idxs.append(-1)
        else:
            idxs.append(idx)

    ch_arg = (args.get("channels") or "").strip()
    selected_list = [c.strip() for c in ch_arg.split(",") if c.strip()]
    selected = set(selected_list)

    def resolve_for_selection(key: str) -> str:
        matches = []
        if key in cols:
            matches.append(key)
        suf = f"-{key}"
        for c in cols:
            if c.endswith(suf) and c not in matches:
                matches.append(c)
        if not matches:
            return ""
        if selected:
            for pref in ("A-", "C-"):
                for c in matches:
                    if c in selected and c.startswith(pref):
                        return c
            for c in matches:
                if c in selected:
                    return c
            return ""
        for pref in ("A-", "C-"):
            for c in matches:
                if c.startswith(pref):
                    return c
        return matches[0]

    key_to_col = {
        "Pc": 4, "Pe": 5, "T-sie": 6, "UR-sie": 7,
        "Tc": 8, "Te": 9, "T1": 10, "T2": 11,
        "T3": 12, "T4": 13, "T5": 14, "T6": 15,
        "T7": 16, "I": 17, "F": 18, "V": 19, "W": 20,
    }

    key_to_code = {k: resolve_for_selection(k) for k in key_to_col.keys()}
    fixed_codes = set([v for v in key_to_code.values() if v])

    candidate_codes = [c for c in selected_list if c in cols] if selected_list else list(cols)

    try:
        include_extra = int(str(args.get("include_extra", "1")).strip() or "1")
    except Exception:
        include_extra = 1

    extra_codes = [c for c in candidate_codes if c and c not in fixed_codes]
    if include_extra <= 0:
        extra_codes = []

    def _nat_key(s: str):
        import re as _re
        parts = _re.split(r'(\d+)', s or '')
        out = []
        for part in parts:
            if part.isdigit():
                try:
                    out.append(int(part))
                except Exception:
                    out.append(part)
            else:
                out.append(part.casefold())
        return out

    def _display_name(code: str) -> str:
        ch = channels.get(code)
        nm = ''
        try:
            nm = (ch.name or '').strip() if ch else ''
        except Exception:
            nm = ''
        if nm:
            return nm
        lb = ''
        try:
            lb = (ch.label or '').strip() if ch else ''
        except Exception:
            lb = ''
        return lb or code

    def _template_header(code: str) -> str:
        ch = channels.get(code)
        nm = _display_name(code)
        unit = ''
        try:
            unit = (ch.unit or '').strip() if ch else ''
        except Exception:
            unit = ''
        if unit:
            return f"{nm} [{unit}]"
        return nm

    extra_codes.sort(key=lambda c: (_nat_key(_display_name(c)), _nat_key(c)))

    EXTRA_START_COL = 26
    extra_col_map = {code: (EXTRA_START_COL + i) for i, code in enumerate(extra_codes)}

    template_path = TEMPLATE_FILE
    if not os.path.isfile(template_path):
        return jsonify({"ok": False, "error": "Не найден template.xlsx"}), 400

    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.formula.translate import Translator
    from copy import copy
    import datetime as _dt

    viewer_settings = get_viewer_settings()

    t_all0 = time_mod.perf_counter()

    wb = load_workbook(template_path, data_only=False, keep_vba=False)
    ws = wb.active

    # refrigerant B1
    refrigerant = (args.get('refrigerant') or 'R290').strip()
    if refrigerant not in ('R290', 'R600a'):
        refrigerant = 'R290'
    try:
        ws['B1'].value = refrigerant
    except Exception:
        pass

    # test folder path D1
    try:
        ws['D1'].value = (STATE.get('folder') or '').strip()
    except Exception:
        pass

    try:
        wb.calculation.calcMode = 'manual'
    except Exception:
        pass

    t_load = time_mod.perf_counter()

    start_row = 4
    pattern_row = start_row

    required_last_col = 0
    if extra_codes:
        required_last_col = EXTRA_START_COL + len(extra_codes) - 1
    BASE_LAST_COL = 25
    max_col = max(BASE_LAST_COL, required_last_col or 0)

    # formula columns and pattern formulas
    formula_cols = set()
    pattern_formulas = {}
    for c in range(1, max_col + 1):
        v = ws.cell(row=pattern_row, column=c).value
        if isinstance(v, str) and v.startswith('='):
            formula_cols.add(c)
            pattern_formulas[c] = v
    formula_list = sorted(list(formula_cols))

    translators = {}
    for c, src_formula in pattern_formulas.items():
        src_addr = f"{get_column_letter(c)}{pattern_row}"
        try:
            translators[c] = Translator(src_formula, origin=src_addr)
        except Exception:
            translators[c] = None

    writers = []
    for key, colnum in key_to_col.items():
        code = key_to_code.get(key) or ""
        if not code:
            continue
        if selected and code not in selected:
            continue
        writers.append((code, colnum))

    extra_writers = []
    if extra_codes:
        for code in extra_codes:
            colnum = extra_col_map.get(code)
            if colnum:
                extra_writers.append((code, colnum))

    base_raw_cols = sorted(set(key_to_col.values()))

    needed_last_row = start_row + len(idxs) - 1

    template_max_row = ws.max_row
    if needed_last_row > template_max_row:
        style_cache = {}
        for c in range(1, max_col + 1):
            src = ws.cell(row=pattern_row, column=c)
            try:
                style_cache[c] = (copy(src._style), src.number_format)
            except Exception:
                style_cache[c] = (None, src.number_format)

        pat_height = None
        try:
            pat_height = ws.row_dimensions[pattern_row].height
        except Exception:
            pat_height = None

        for r in range(template_max_row + 1, needed_last_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                st, nf = style_cache.get(c, (None, None))
                if st is not None:
                    cell._style = st
                if nf is not None:
                    cell.number_format = nf
            if pat_height is not None:
                try:
                    ws.row_dimensions[r].height = pat_height
                except Exception:
                    pass

    t_prep = time_mod.perf_counter()

    for j, idx in enumerate(idxs):
        r = start_row + j

        ws.cell(row=r, column=2).value = _dt.timedelta(seconds=j * 20)

        if idx >= 0:
            tms = rows[idx]["t_ms"]
            dt_obj = dt.datetime.fromtimestamp(tms / 1000)
            ws.cell(row=r, column=3).value = dt_obj.time()
        else:
            ws.cell(row=r, column=3).value = None

        for colnum in base_raw_cols:
            ws.cell(row=r, column=colnum).value = None

        if idx >= 0:
            row = rows[idx]
            for code, colnum in writers:
                v = row.get(code)
                if v is not None:
                    try:
                        ws.cell(row=r, column=colnum).value = float(v)
                    except Exception:
                        pass

            for code, colnum in extra_writers:
                v = row.get(code)
                if v is None:
                    ws.cell(row=r, column=colnum).value = None
                else:
                    try:
                        ws.cell(row=r, column=colnum).value = float(v)
                    except Exception:
                        ws.cell(row=r, column=colnum).value = None
        else:
            for _, colnum in extra_writers:
                ws.cell(row=r, column=colnum).value = None

        for c in formula_list:
            cell = ws.cell(row=r, column=c)
            vv = cell.value
            if isinstance(vv, str) and vv.startswith('='):
                continue

            src_formula = pattern_formulas.get(c)
            if not src_formula:
                continue

            tr = translators.get(c)
            if tr is not None:
                dst_addr = f"{get_column_letter(c)}{r}"
                try:
                    cell.value = tr.translate_formula(dst_addr)
                except Exception:
                    cell.value = src_formula
            else:
                cell.value = src_formula

    if extra_writers:
        header_row = start_row - 1
        for code, colnum in extra_writers:
            try:
                ws.cell(row=header_row, column=colnum).value = _template_header(code)
            except Exception:
                ws.cell(row=header_row, column=colnum).value = _display_name(code)

    clear_from = start_row + len(idxs)
    clear_to = min(ws.max_row, clear_from + 200)
    clear_cols = [2, 3] + base_raw_cols + [c for _, c in extra_writers]
    cells_dict = getattr(ws, '_cells', None)

    for r in range(clear_from, clear_to + 1):
        for c in clear_cols:
            if cells_dict is not None:
                cell = cells_dict.get((r, c))
                if cell is not None:
                    cell.value = None
            else:
                ws.cell(row=r, column=c).value = None

    t_write = time_mod.perf_counter()

    # conditional formatting
    try:
        from openpyxl.styles import PatternFill
        from openpyxl.formatting.rule import FormulaRule

        first_r = start_row
        last_r = start_row + len(idxs) - 1

        if last_r >= first_r:
            rm = viewer_settings.get('row_mark') if isinstance(viewer_settings.get('row_mark'), dict) else {}
            thr = rm.get('threshold_T', 150)
            try:
                thr_f = float(thr)
                thr = int(thr_f) if thr_f.is_integer() else thr_f
            except Exception:
                thr = 150

            color_hex = str(rm.get('color') or '#EAD706')
            intensity = rm.get('intensity', 100)
            try:
                intensity = int(intensity)
            except Exception:
                intensity = 100

            fill_argb = _argb_from_hex_and_intensity(color_hex, intensity)
            row_fill = PatternFill(fill_type='solid', start_color=fill_argb, end_color=fill_argb)

            rng_row = f"B{first_r}:P{last_r}"
            rule_row = FormulaRule(formula=[f"$T{first_r}<{thr}"], fill=row_fill, stopIfTrue=True)
            try:
                rule_row.priority = 1
            except Exception:
                pass
            ws.conditional_formatting.add(rng_row, rule_row)

            def _fmt_thr(x) -> str | None:
                try:
                    xf = float(x)
                    if xf.is_integer():
                        return str(int(xf))
                    return (f"{xf:g}").replace(',', '.')
                except Exception:
                    return None

            dm = viewer_settings.get('discharge_mark') if isinstance(viewer_settings.get('discharge_mark'), dict) else {}
            td_thr_s = _fmt_thr(dm.get('threshold', None))
            if td_thr_s is not None:
                hx = _normalize_hex_color(str(dm.get('color') or ''), default='#FFC000')
                argb = 'FF' + hx[1:].upper()
                f_td = PatternFill(fill_type='solid', start_color=argb, end_color=argb)
                rng_td = f"H{first_r}:H{last_r}"
                rule_td = FormulaRule(formula=[f'AND($H{first_r}<>"",$H{first_r}>{td_thr_s})'], fill=f_td, stopIfTrue=True)
                try:
                    rule_td.priority = 5
                except Exception:
                    pass
                ws.conditional_formatting.add(rng_td, rule_td)

            sm = viewer_settings.get('suction_mark') if isinstance(viewer_settings.get('suction_mark'), dict) else {}
            ts_thr_s = _fmt_thr(sm.get('threshold', None))
            if ts_thr_s is not None:
                hx = _normalize_hex_color(str(sm.get('color') or ''), default='#00B0F0')
                argb = 'FF' + hx[1:].upper()
                f_ts = PatternFill(fill_type='solid', start_color=argb, end_color=argb)
                rng_ts = f"I{first_r}:I{last_r}"
                rule_ts = FormulaRule(formula=[f'AND($I{first_r}<>"",$I{first_r}<{ts_thr_s})'], fill=f_ts, stopIfTrue=True)
                try:
                    rule_ts.priority = 6
                except Exception:
                    pass
                ws.conditional_formatting.add(rng_ts, rule_ts)

            scales = viewer_settings.get('scales') if isinstance(viewer_settings.get('scales'), dict) else {}

            def _hex_to_argb(hx: str, default: str) -> str:
                hx = _normalize_hex_color(str(hx or ''), default=default)
                return 'FF' + hx[1:].upper()

            def _fmt_num(x) -> str:
                try:
                    xf = float(x)
                    if xf.is_integer():
                        return str(int(xf))
                    return (f"{xf:g}").replace(',', '.')
                except Exception:
                    return str(x)

            def _add_discrete(col_letter: str, base_prio: int) -> None:
                spec = scales.get(col_letter) if isinstance(scales.get(col_letter), dict) else {}
                dflt = DEFAULT_VIEWER_SETTINGS['scales'][col_letter]

                vmin = spec.get('min', dflt.get('min'))
                vopt = spec.get('opt', dflt.get('opt'))

                try:
                    vmin_f = float(vmin)
                except Exception:
                    vmin_f = float(dflt.get('min') or 0)
                try:
                    vopt_f = float(vopt)
                except Exception:
                    vopt_f = float(dflt.get('opt') or 0)

                if vmin_f > vopt_f:
                    vmin_f, vopt_f = vopt_f, vmin_f

                min_s = _fmt_num(vmin_f)
                opt_s = _fmt_num(vopt_f)

                colors = spec.get('colors') if isinstance(spec.get('colors'), dict) else {}
                dcolors = dflt.get('colors') if isinstance(dflt.get('colors'), dict) else {}

                c_lo = _hex_to_argb(colors.get('min'), dcolors.get('min', '#1CBCF2'))
                c_mid = _hex_to_argb(colors.get('opt'), dcolors.get('opt', '#00FF00'))
                c_hi = _hex_to_argb(colors.get('max'), dcolors.get('max', '#F3919B'))

                f_lo = PatternFill(fill_type='solid', start_color=c_lo, end_color=c_lo)
                f_mid = PatternFill(fill_type='solid', start_color=c_mid, end_color=c_mid)
                f_hi = PatternFill(fill_type='solid', start_color=c_hi, end_color=c_hi)

                rng = f"{col_letter}{first_r}:{col_letter}{last_r}"

                r1 = FormulaRule(formula=[f'AND(${col_letter}{first_r}<>"",${col_letter}{first_r}<{min_s})'], fill=f_lo, stopIfTrue=True)
                r2 = FormulaRule(formula=[f'AND(${col_letter}{first_r}<>"",${col_letter}{first_r}>={min_s},${col_letter}{first_r}<={opt_s})'], fill=f_mid, stopIfTrue=True)
                r3 = FormulaRule(formula=[f'AND(${col_letter}{first_r}<>"",${col_letter}{first_r}>{opt_s})'], fill=f_hi, stopIfTrue=True)

                try:
                    r1.priority = base_prio
                    r2.priority = base_prio + 1
                    r3.priority = base_prio + 2
                except Exception:
                    pass

                ws.conditional_formatting.add(rng, r1)
                ws.conditional_formatting.add(rng, r2)
                ws.conditional_formatting.add(rng, r3)

            _add_discrete('W', 10)
            _add_discrete('X', 20)
            _add_discrete('Y', 30)

    except Exception as _cf_e:
        try:
            print('[TEMPLATE] conditional formatting skipped:', _cf_e)
        except Exception:
            pass

    t_before_save = time_mod.perf_counter()

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    t_after_save = time_mod.perf_counter()

    total_s = t_after_save - t_all0
    load_s = t_load - t_all0
    prep_s = t_prep - t_load
    write_s = t_write - t_prep
    format_s = t_before_save - t_write
    save_s = t_after_save - t_before_save

    try:
        print(
            f"[TEMPLATE] timing: load={load_s:.3f}s prep={prep_s:.3f}s write={write_s:.3f}s format={format_s:.3f}s save={save_s:.3f}s total={total_s:.3f}s rows={len(idxs)} max_col={max_col}")
    except Exception:
        pass

    fn = f"template_filled_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    # NOTE: send_file_compat is applied by route layer (needs Flask send_file)
    return bio, fn, {
        'X-Export-Total-S': f"{total_s:.3f}",
        'X-Export-Timing': f"load {load_s:.1f}s | prep {prep_s:.1f}s | write {write_s:.1f}s | format {format_s:.1f}s | save {save_s:.1f}s | total {total_s:.1f}s",
    }
